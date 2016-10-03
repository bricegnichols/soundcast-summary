
# coding: utf-8

# In[11]:

import pandas as pd
import numpy as np
import plotly
import cufflinks as cf
from plotly.graph_objs import Scatter, Layout
import plotly.graph_objs as go
import summary_functions as scf
import plotly.plotly as py


py.sign_in('paulsameyt', 'qm6nrkka4z') # Replace the username, and API key with your credentials.

# In[12]:

plotly.offline.init_notebook_mode()
cf.set_config_file(offline=True, world_readable=True, theme='ggplot')


# In[13]:

input_file = 'network_summary_detailed.xlsx'


# In[14]:

df = pd.read_excel('network_summary_detailed.xlsx','Network Summary')
df_tidy = pd.melt(df,
                  id_vars=['tod','TP_4k'],
                  var_name = 'category',
                  value_name = 'Model Run')
df_split = pd.DataFrame(df_tidy.category.str.split('_',1).tolist(),
                        columns=['road_type','table_type'])
df_cat = pd.concat([df_tidy,df_split],axis=1)


# In[15]:

df_cat.rename(columns={'TP_4k':'Time'},inplace=True)


# In[16]:

df_cat_vmt = df_cat[df_cat['table_type']=='vmt'] 
df_cat_vht = df_cat[df_cat['table_type']=='vht'] 
df_cat_delay = df_cat[df_cat['table_type']=='delay'] 


# In[17]:

df_cat_vmt_byTime = df_cat_vmt.groupby('Time').sum().reindex(['am','md','pm','ev','ni'])
df_cat_vmt_byRoad = df_cat_vmt.groupby('road_type').sum().reindex(['highway','arterial','connectors'])
df_cat_vht_byTime = df_cat_vht.groupby('Time').sum().reindex(['am','md','pm','ev','ni'])
df_cat_vht_byRoad = df_cat_vht.groupby('road_type').sum().reindex(['highway','arterial','connectors'])
df_cat_delay_byTime = df_cat_delay.groupby('Time').sum().reindex(['am','md','pm','ev','ni'])
df_cat_delay_byRoad = df_cat_delay.groupby('road_type').sum().reindex(['highway','arterial','connectors'])


# In[18]:

from openpyxl import load_workbook
wb = load_workbook(filename = 'NetworkSummaryTemplate.xlsx')
ws = wb['Network']


# In[19]:

# Summary plots
df_cat_vmt_byTime['Comparison Scenario'] = [int(ws['B'+str(3+i)].value) for i in range(len(df_cat_vmt_byTime))]
trace1 = go.Bar(x=df_cat_vmt_byTime.index,y=df_cat_vmt_byTime['Model Run'],
                name='Model Run')
trace2 = go.Bar(x=df_cat_vmt_byTime.index,y=df_cat_vmt_byTime['Comparison Scenario'],
                name='Comparison Scenario')
data = [trace1,trace2]
layout = go.Layout(title='vmt by Time',barmode='group',xaxis = dict(title ='Time'),yaxis = dict(title ='Number of Vehicles'))
fig = go.Figure(data=data, layout=layout)
py.image.save_as(fig, filename='vmt_by_Time.png')
# df_cat_vmt_byTime.iplot(kind='bar',title='vmt by Time')



# In[20]:

df_cat_vmt_byRoad['Comparison Scenario'] = [int(ws['B'+str(10+i)].value) for i in range(len(df_cat_vmt_byRoad))]
trace1 = go.Bar(x=df_cat_vmt_byRoad.index,y=df_cat_vmt_byRoad['Model Run'],
                name='Model Run')
trace2 = go.Bar(x=df_cat_vmt_byRoad.index,y=df_cat_vmt_byRoad['Comparison Scenario'],
                name='Comparison Scenario')
data = [trace1,trace2]
layout = go.Layout(title='vmt by Roadtype',barmode='group',xaxis = dict(title ='Road Type'),yaxis = dict(title ='Number of Vehicles'))
fig = go.Figure(data=data, layout=layout)
py.image.save_as(fig, filename='vmt_by_Roadtype.png')
# df_cat_vmt_byRoad.iplot(kind='bar',title='vmt by roadtype')


# In[21]:

df_cat_vht_byTime['Comparison Scenario'] = [int(ws['B'+str(17+i)].value) for i in range(len(df_cat_vht_byTime))]
trace1 = go.Bar(x=df_cat_vht_byTime.index,y=df_cat_vht_byTime['Model Run'],
                name='Model Run')
trace2 = go.Bar(x=df_cat_vht_byTime.index,y=df_cat_vht_byTime['Comparison Scenario'],
                name='Comparison Scenario')
data = [trace1,trace2]
layout = go.Layout(title='vht by Time',barmode='group',xaxis = dict(title ='Time'),yaxis = dict(title ='Number of Vehicles'))
fig = go.Figure(data=data, layout=layout)
py.image.save_as(fig, filename='vht_by_Time.png')
# df_cat_vht_byTime.iplot(kind='bar',title='vht by Time')


# In[22]:

df_cat_vht_byRoad['Comparison Scenario'] = [int(ws['B'+str(24+i)].value) for i in range(len(df_cat_vht_byRoad))]
trace1 = go.Bar(x=df_cat_vht_byRoad.index,y=df_cat_vht_byRoad['Model Run'],
                name='Model Run')
trace2 = go.Bar(x=df_cat_vht_byRoad.index,y=df_cat_vht_byRoad['Comparison Scenario'],
                name='Comparison Scenario')
data = [trace1,trace2]
layout = go.Layout(title='vht by Roadtype',barmode='group',xaxis = dict(title ='Road Type'),yaxis = dict(title ='Number of Vehicles'))
fig = go.Figure(data=data, layout=layout)
py.image.save_as(fig, filename='vht_by_Roadtype.png')
# df_cat_vht_byRoad.iplot(kind='bar',title='vht by Road')


# In[23]:

df_cat_delay_byTime['Comparison Scenario'] = [int(ws['B'+str(31+i)].value) for i in range(len(df_cat_delay_byTime))]
trace1 = go.Bar(x=df_cat_delay_byTime.index,y=df_cat_delay_byTime['Model Run'],
                name='Model Run')
trace2 = go.Bar(x=df_cat_delay_byTime.index,y=df_cat_delay_byTime['Comparison Scenario'],
                name='Comparison Scenario')
data = [trace1,trace2]
layout = go.Layout(title='delay by Time',barmode='group',xaxis = dict(title ='Time'),yaxis = dict(title ='Number of Vehicles'))
fig = go.Figure(data=data, layout=layout)
py.image.save_as(fig, filename='delay_by_Time.png')
# df_cat_delay_byTime.iplot(kind='bar',title='delay by Time')


# In[24]:

df_cat_delay_byRoad['Comparison Scenario'] = [int(ws['B'+str(38+i)].value) for i in range(len(df_cat_delay_byRoad))]
trace1 = go.Bar(x=df_cat_delay_byRoad.index,y=df_cat_delay_byRoad['Model Run'],
                name='Model Run')
trace2 = go.Bar(x=df_cat_delay_byRoad.index,y=df_cat_delay_byRoad['Comparison Scenario'],
                name='Comparison Scenario')
data = [trace1,trace2]
layout = go.Layout(title='delay by Roadtype',barmode='group',xaxis = dict(title ='Road Type'),yaxis = dict(title ='Number of Vehicles'))
fig = go.Figure(data=data, layout=layout)
py.image.save_as(fig, filename='delay_by_Roadtype.png')
# df_cat_delay_byRoad.iplot(kind='bar',title='delay by Road')


# In[25]:

# Count Time Plot
counts_by_tod = pd.DataFrame(columns = ['Counts (Model Run)', 'Counts (Observed)'],
index = ['5 to 6', '6 to 7', '7 to 8', '8 to 9',
'9 to 10', '10 to 14', '14 to 15', '15 to 16',
'16 to 17', '17 to 18', '18 to 20', '20 to 5'])
counts_output = pd.read_excel('network_summary_detailed.xlsx','Counts Output')


# In[26]:

for tod in counts_by_tod.index:
    counts_by_tod.loc[tod, 'Counts (Observed)'] = scf.get_counts(counts_output, tod)
    if tod == '20 to 5':
        counts_by_tod.loc[tod, 'Counts (Observed)'] = counts_by_tod.loc[tod, 'Counts (Observed)'] - counts_by_tod.loc['5 to 6', 'Counts (Observed)']   
    counts_by_tod.loc[tod, 'Counts (Model Run)'] = counts_output['vol' + tod.replace(' ', '')].sum()


# In[27]:
trace1 = go.Scatter(x=counts_by_tod.index,y=counts_by_tod['Counts (Model Run)'],name='Model Run')
trace2 = go.Scatter(x=counts_by_tod.index,y=counts_by_tod['Counts (Observed)'],name='Observed')
data = [trace1, trace2]
layout = dict(title = 'Time of Day',
              xaxis = dict(title = 'Time of Day'),
              yaxis = dict(title = 'Number of Vehicles'),
              )
fig = go.Figure(data=data,layout=layout)
py.image.save_as(fig,filename='time_of_Day.png')

# counts_by_tod.iplot(kind='line',title='Time of Day',xTitle='Time of Day',yTitle='Number of Vehicles')
# py.image.save_as(fig, filename='timeOfDay.png')


# In[28]:

# Count all plot
counts_all = pd.read_excel(input_file,sheetname='Counts Output')
counts_all = counts_all.reset_index()
counts_all = counts_all.fillna(0)
counts_all['Total'] = counts_all['vol5to6'] + counts_all['vol6to7'] + counts_all['vol7to8'] + counts_all['vol8to9'] + counts_all['vol9to10'] + counts_all['vol10to14'] + counts_all['vol14to15'] + counts_all['vol15to16'] + counts_all['vol16to17'] + counts_all['vol17to18'] + counts_all['vol18to20'] + counts_all['vol20to5']

r2 = (counts_all[['Vol_Daily','Total']].corr() **2).loc['Vol_Daily','Total']
slope = (counts_all[['Vol_Daily','Total']].cov()).loc['Vol_Daily','Total']/counts_all['Vol_Daily'].var()
intercept = counts_all['Total'].mean() - slope * counts_all['Vol_Daily'].mean()
columns = counts_all.columns.tolist()
index = counts_all.index.tolist()
xi = np.array(counts_all['Vol_Daily'])
line = slope * xi + intercept
y = np.array(counts_all['Total'])


# In[29]:

# Creating the dataset, and generating the plot
trace1 = go.Scatter(
                  x=xi, 
                  y=y, 
                  mode='markers',
                  marker=go.Marker(color='rgb(255, 127, 14)'),
                  name='Data'
                  )

trace2 = go.Scatter(
                  x=xi, 
                  y=line, 
                  mode='lines',
                  marker=go.Marker(color='rgb(31, 119, 180)'),
                  name='Fit'
                  )

annotation = go.Annotation(
                  x=40000,
                  y = 120000,
                  text='R^2 = ' + str(round(r2,3)) + ', Y =' + str(round(slope,3)) + 'X + ' + str(round(intercept,3)),
                  showarrow=False,
                  font=go.Font(size=16)
                  )
layout = go.Layout(
                title='Modeled vs. Observed Counts',
                plot_bgcolor='rgb(229, 229, 229)',
                  xaxis=dict(title = 'Observed Counts',zeroline= False,zerolinecolor='rgb(255,255,255)',gridcolor='rgb(255,255,255)'),
                  yaxis=dict(title = 'Modeled Counts',zeroline= False,zerolinecolor='rgb(255,255,255)', gridcolor='rgb(255,255,255)'),
                  annotations=[annotation]
                )


# In[30]:

data = [trace1, trace2]
fig = go.Figure(data=data, layout=layout)
py.image.save_as(fig, filename='data_fit.png')
# plotly.offline.iplot(fig)


# In[10]: For screeline plots

screenline_dict = {'Primary': {
4: 'Tacoma - East of CBD',
14: 'Auburn',
15: 'Auburn',
22: 'Tukwila',
23: 'Renton',
29: 'Seattle - South of CBD',
30: 'Bellevue/Redmond',
32: 'TransLake',
35: 'Ship Canal',
37: 'Kirkland/Redmond',
41: 'Seattle - North',
43: 'Lynnwood/Bothell',
44: 'Bothell',
46: 'Mill Creek'},
'Secondary': {
2: 'Parkland',
3: 'Puyallup',
7: 'Tacoma Narrows',
18: 'Maple Valley',
19: 'SeaTac',
20: 'Kent',
54: 'Gig Harbor',
57: 'Kitsap - North',
58: 'Agate Pass',
60: 'Cross-Sound',
66: 'Preston, Issaquah',
71: 'Woodinville'}}

observed_screenline_volumes = {'Tacoma - East of CBD': 271777,
'Auburn': 534811,
'Tukwila': 239527,
'Renton': 81758,
'Seattle - South of CBD': 490806,
'Bellevue/Redmond': 354612,
'TransLake': 250220,
'Ship Canal': 521155,
'Kirkland/Redmond': 381331,
'Seattle - North': 327021,
'Lynnwood/Bothell': 231368,
'Bothell': 255590,
'Mill Creek': 350492,
'Parkland': 285859,
'Puyallup': 118726,
'Tacoma Narrows': 79000,
'Maple Valley': 61921,
'SeaTac': 71364,
'Kent': 504607,
'Gig Harbor': 58503,
'Kitsap - North': 97177,
'Agate Pass': 21000,
'Cross-Sound': 17466,
'Preston, Issaquah': 93227,
'Woodinville': 98331}


# In[11]:

# primary screenline plot
screenline_df = pd.read_excel(input_file,sheetname='Screenline Volumes')
screenline_type = 'Primary'
screenline_df['Screenline Name'] = screenline_df['Screenline'].map(screenline_dict[screenline_type])
screenline_df = screenline_df.groupby('Screenline Name').sum()
screenline_df.loc['Auburn', 'Screenline'] = '14/15'
screenline_df = screenline_df.dropna()
screenline_df = screenline_df.reset_index()
screenline_df['Observed Volume'] = screenline_df['Screenline Name'].map(observed_screenline_volumes)
screenline_df = screenline_df.set_index('Screenline Name')
screenline_df['Modeled Volume'] = screenline_df['Volumes']
del screenline_df['Volumes']
screenline_df = screenline_df[['Screenline', 'Modeled Volume', 'Observed Volume']]
screenline_df['Est/Obs'] = (screenline_df['Modeled Volume']/screenline_df['Observed Volume']).round(2)
screenline_df_primary = scf.get_differences(screenline_df, 'Modeled Volume', 'Observed Volume', -2)
# screenline_df_primary


# In[12]:
trace1 = go.Bar(x=screenline_df_primary.index,y=screenline_df_primary['Modeled Volume'],
                name='Modeled Volume')
trace2 = go.Bar(x=screenline_df_primary.index,y=screenline_df_primary['Observed Volume'],
                name='Observed Volume')
data = [trace1,trace2]
layout = go.Layout(title='Model vs. Observed counts by Primary Screenline',barmode='group',
  xaxis = dict(title ='Primary Screeline'),yaxis = dict(title ='Counts'))
fig = go.Figure(data=data, layout=layout)
py.image.save_as(fig, filename='Primary_Screenline.png')

# screenline_df_primary[['Modeled Volume','Observed Volume']].iplot(kind='bar',title='Model vs. Observed counts by Primary Screenline',
#                                                                   xTitle='Primary Screeline',yTitle='Counts')


# In[13]:

# secondary screenline plot
screenline_df = pd.read_excel(input_file,sheetname='Screenline Volumes')
screenline_type = 'Secondary'
screenline_df['Screenline Name'] = screenline_df['Screenline'].map(screenline_dict[screenline_type])
screenline_df = screenline_df.groupby('Screenline Name').sum()
screenline_df.loc['Auburn', 'Screenline'] = '14/15'
screenline_df = screenline_df.dropna()
screenline_df = screenline_df.reset_index()
screenline_df['Observed Volume'] = screenline_df['Screenline Name'].map(observed_screenline_volumes)
screenline_df = screenline_df.set_index('Screenline Name')
screenline_df['Modeled Volume'] = screenline_df['Volumes']
del screenline_df['Volumes']
screenline_df = screenline_df[['Screenline', 'Modeled Volume', 'Observed Volume']]
screenline_df['Est/Obs'] = (screenline_df['Modeled Volume']/screenline_df['Observed Volume']).round(2)
screenline_df_secondary = scf.get_differences(screenline_df, 'Modeled Volume', 'Observed Volume', -2)
# screenline_df_secondary


# In[14]:
trace1 = go.Bar(x=screenline_df_secondary.index,y=screenline_df_secondary['Modeled Volume'],
                name='Modeled Volume')
trace2 = go.Bar(x=screenline_df_secondary.index,y=screenline_df_secondary['Observed Volume'],
                name='Observed Volume')
data = [trace1,trace2]
layout = go.Layout(title='Model vs. Observed counts by Secondary Screenline',barmode='group',
  xaxis = dict(title ='Secondary Screeline'),yaxis = dict(title ='Counts'))
fig = go.Figure(data=data, layout=layout)
py.image.save_as(fig, filename='Secondary_Screenline.png')

# screenline_df_secondary[['Modeled Volume','Observed Volume']].iplot(kind='bar',title='Model vs. Observed counts by Secondary Screenline',
#                                                                   xTitle='Secondary Screeline',yTitle='Counts')




